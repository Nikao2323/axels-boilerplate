"""
DWG vs Excel Volume Comparison Tool — Phase 2
გამოყენება: python3 compare.py <dwg_or_dxf_file> <excel_file>
"""

import sys
import os
import shutil
import subprocess
import tempfile
import openpyxl
import ezdxf


AUTOCAD_WSL = os.environ.get(
    "ARSI_AUTOCAD_EXE",
    "/mnt/c/Program Files/Autodesk/AutoCAD 2021/accoreconsole.exe",
)

# Calibrated from 6 reference buildings (სახლი 6,8,9,11,12,13)
RATIO_CHECKS = [
    {
        "key":      "rkinabetonis karkasi",
        "unit":     "m2",
        "label":    "კარკასი",
        "min":      1.22,
        "max":      1.27,
        "tol":      0.05,    # ±5% extra tolerance
        "warn_only": False,
    },
    {
        "key":      "betoni b-25",
        "unit":     "m3",
        "label":    "ბეტონი B-25",
        "min":      0.276,
        "max":      0.315,
        "tol":      0.05,
        "warn_only": False,
    },
    {
        "key":      "armatura a500",
        "unit":     "t",
        "label":    "არმატ. A500",
        "min":      0.035,
        "max":      0.042,
        "tol":      0.05,
        "warn_only": False,
    },
    {
        "key":      "qvabulis amoReba",
        "unit":     "m3",
        "label":    "ქვაბ. ამოღება",
        "min":      0.62,
        "max":      0.82,
        "tol":      0.08,    # wider — more variable
        "warn_only": True,   # only warning, not failure
    },
]

TOLERANCE_AREA_PCT = 8   # % floor area match tolerance (DXF vs Excel)


# ── DWG → DXF conversion ─────────────────────────────────────────────────────

def dwg_to_dxf(dwg_path: str) -> str:
    if dwg_path.lower().endswith(".dxf"):
        return dwg_path

    # Fixed Windows-accessible staging paths (no deep temp subdirs)
    WIN_TEMP_WSL = "/mnt/c/Users/Nika/AppData/Local/Temp"

    staged_dwg_wsl = f"{WIN_TEMP_WSL}/staged_input.dwg"
    staged_dxf_wsl = f"{WIN_TEMP_WSL}/staged_output.dxf"
    staged_dwg_win = "C:\\Users\\Nika\\AppData\\Local\\Temp\\staged_input.dwg"
    staged_dxf_win = "C:\\Users\\Nika\\AppData\\Local\\Temp\\staged_output.dxf"
    scr_wsl        = f"{WIN_TEMP_WSL}/staged_convert.scr"
    scr_win        = "C:\\Users\\Nika\\AppData\\Local\\Temp\\staged_convert.scr"

    # Copy input DWG to known Windows-accessible path
    shutil.copy2(dwg_path, staged_dwg_wsl)

    # Remove stale output
    if os.path.exists(staged_dxf_wsl):
        os.remove(staged_dxf_wsl)

    # Write AutoCAD script
    with open(scr_wsl, "w", encoding="utf-8") as f:
        f.write(f"DXFOUT\n{staged_dxf_win}\n16\n\n")

    # Launch accoreconsole and poll for output file (don't rely on process exit)
    proc = subprocess.Popen(
        [AUTOCAD_WSL, "/i", staged_dwg_win, "/s", scr_win, "/l", "en-US"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    import time
    deadline = time.time() + 300  # 5 min max
    while time.time() < deadline:
        if os.path.exists(staged_dxf_wsl) and os.path.getsize(staged_dxf_wsl) > 1000:
            time.sleep(1)  # let AutoCAD finish writing
            proc.terminate()
            return staged_dxf_wsl
        time.sleep(2)

    proc.terminate()
    raise FileNotFoundError(f"DXF conversion timed out: {dwg_path}")


# ── DXF geometry extraction ───────────────────────────────────────────────────

def _poly_area(pts):
    x = [p[0] for p in pts]
    y = [p[1] for p in pts]
    n = len(x)
    return abs(sum(x[i]*y[(i+1)%n] - x[(i+1)%n]*y[i] for i in range(n))) / 2


def extract_floor_polygons(dxf_path: str, min_area=200, max_area=3000):
    """Return clustered floor-polygon list: [{area, count}, ...]  (m², descending)."""
    doc = ezdxf.readfile(dxf_path)
    raw = []
    for ent in doc.modelspace():
        if ent.dxftype() == "LWPOLYLINE" and ent.is_closed:
            pts = list(ent.vertices())
            if len(pts) >= 4:
                a = _poly_area(pts)
                if min_area < a < max_area:
                    raw.append(round(a, 1))

    raw.sort()
    clusters = []
    for a in raw:
        if clusters and abs(a - clusters[-1][0]) < 3:
            clusters[-1].append(a)
        else:
            clusters.append([a])

    result = []
    for c in clusters:
        rep = sorted(c)[len(c) // 2]
        result.append({"area": rep, "count": len(c)})
    return sorted(result, key=lambda x: -x["area"])


# ── Excel extraction ──────────────────────────────────────────────────────────

def extract_excel_floors(xlsx_path: str) -> dict:
    """Return {label: area_m2} from ფართები sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "ფართები" not in wb.sheetnames:
        return {}
    floors = {}
    for row in wb["ფართები"].iter_rows(min_row=4, values_only=True):
        if len(row) < 3:
            continue
        _, floor, total = row[0], row[1], row[2]
        if not (floor and isinstance(total, (int, float)) and total > 0):
            continue
        s = str(floor)
        if "jami" in s.lower():
            floors["სულ"] = float(total)
        elif "Zirk" in s or "zirk" in s:
            floors["საძირკველი"] = float(total)
        elif "xuravi" in s or "saxuravi" in s:
            floors["სახურავი"] = float(total)
        elif isinstance(floor, (int, float)):
            floors[f"სართ.{int(floor)}"] = float(total)
    return floors


def extract_excel_budget(xlsx_path: str) -> list:
    """Return [{desc, unit, qty}] from ბიუჯეტი sheet (qty > 10, ASCII units only)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "ბიუჯეტი" not in wb.sheetnames:
        return []
    items = []
    header = False
    for row in wb["ბიუჯეტი"].iter_rows(min_row=1, values_only=True):
        if not header:
            if row[1] == "samuSaos dasaxeleba" or row[0] == "#":
                header = True
            continue
        desc, unit, qty = row[1], row[2], row[3]
        if (isinstance(desc, str)
                and isinstance(qty, (int, float))
                and qty > 10
                and isinstance(unit, str)
                and unit in ("m2", "m3", "t")):
            items.append({"desc": desc, "unit": unit, "qty": float(qty)})
    return items


# ── Matching helpers ──────────────────────────────────────────────────────────

def _best_polygon(excel_area, dxf_polys):
    if not dxf_polys:
        return None, None, None
    best = min(dxf_polys, key=lambda p: abs(p["area"] - excel_area))
    diff_pct = (best["area"] - excel_area) / excel_area * 100
    ok = abs(diff_pct) <= TOLERANCE_AREA_PCT
    return best["area"], diff_pct, ok


def _find_budget_item(key, unit, budget):
    """Return the first matching budget item with given key substring and unit."""
    key_lower = key.lower()
    for item in budget:
        if key_lower in item["desc"].lower() and item["unit"] == unit:
            return item
    return None


def _ratio_status(qty, total, r):
    lo  = total * r["min"] * (1 - r["tol"])
    hi  = total * r["max"] * (1 + r["tol"])
    exp = f"{total*r['min']:.0f}–{total*r['max']:.0f}"
    if lo <= qty <= hi:
        return True, f"✓  ({exp} {r['unit']})"
    else:
        tag = "⚠" if r["warn_only"] else "✗"
        return r["warn_only"], f"{tag} მოსალოდნელი {exp} {r['unit']}, მიღებული {qty:.1f}"


# ── Main report ───────────────────────────────────────────────────────────────

def compare(dwg_path: str, xlsx_path: str):
    print(f"\n{'='*68}")
    print(f"  DWG  : {os.path.basename(dwg_path)}")
    print(f"  Excel: {os.path.basename(xlsx_path)}")
    print("="*68)

    print("\n▶ DWG → DXF კონვერტაცია...")
    dxf_path = dwg_to_dxf(dwg_path)
    print(f"  OK: {os.path.basename(dxf_path)}")

    dxf_polys  = extract_floor_polygons(dxf_path)
    xl_floors  = extract_excel_floors(xlsx_path)
    xl_budget  = extract_excel_budget(xlsx_path)
    total_xl   = xl_floors.get("სულ")

    failures = []
    warnings = []

    # ── A: Floor areas ────────────────────────────────────────────────────
    print("\n── A. სართულების ფართი (Excel ფართები vs DWG polygons) " + "─"*15)
    print(f"  {'სართული':<14} {'Excel':>8}  {'DWG':>8}  {'სხვაობა':>8}  სტატ.")
    print("  " + "─"*55)

    dxf_sum = 0
    for label, ea in sorted(xl_floors.items(), key=lambda x: -x[1]):
        if label == "სულ":
            continue
        da, dp, ok = _best_polygon(ea, dxf_polys)
        da_s  = f"{da:.1f}" if da else "—"
        dp_s  = f"{dp:+.1f}%" if dp is not None else "—"
        tag   = "✓" if ok else "✗"
        print(f"  {label:<14} {ea:>8.1f}  {da_s:>8}  {dp_s:>8}  {tag}")
        if da:
            dxf_sum += da
        if not ok:
            failures.append(f"სართ. {label}: Excel {ea:.1f} vs DWG {da_s} ({dp_s})")

    if total_xl:
        dp = (dxf_sum - total_xl) / total_xl * 100
        ok = abs(dp) <= 10
        tag = "✓" if ok else "✗"
        print(f"  {'სულ':<14} {total_xl:>8.1f}  {dxf_sum:>8.1f}  {dp:>+7.1f}%  {tag}")
        if not ok:
            failures.append(f"სულ ფართი: Excel {total_xl:.1f} vs DWG {dxf_sum:.1f} ({dp:+.1f}%)")

    # ── B: Budget volume ratios ───────────────────────────────────────────
    print("\n── B. ბიუჯეტის მოცულობები (raod. სვეტი vs ratio check) " + "─"*14)
    print(f"  {'პუნქტი':<20} {'ერთ.':>4} {'Excel raod.':>12}  შედეგი")
    print("  " + "─"*68)

    if not total_xl:
        print("  ⚠ სულ ფართი ვერ მოიძებნა — ratio check შეუძლებელია")
    else:
        for r in RATIO_CHECKS:
            item = _find_budget_item(r["key"], r["unit"], xl_budget)
            if item is None:
                print(f"  {r['label']:<20} {r['unit']:>4} {'—':>12}  ⚠ ვერ მოიძებნა")
                continue
            qty  = item["qty"]
            ok, msg = _ratio_status(qty, total_xl, r)
            print(f"  {r['label']:<20} {r['unit']:>4} {qty:>12.2f}  {msg}")
            if not ok:
                if r["warn_only"]:
                    warnings.append(f"{r['label']}: {msg}")
                else:
                    failures.append(f"{r['label']}: {msg}")

    # ── C: All budget items table ─────────────────────────────────────────
    print("\n── C. ბიუჯეტის სრული სია (m2/m3/t > 10) " + "─"*27)
    print(f"  {'სამუშაო':<46} {'ერთ.':>4} {'raod.':>12}")
    print("  " + "─"*65)
    for item in xl_budget:
        print(f"  {item['desc'][:46]:<46} {item['unit']:>4} {item['qty']:>12.2f}")

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + "="*68)
    if not failures and not warnings:
        print("  ᲨᲔᲓᲔᲒᲘ: ✓ ყველა შემოწმება გაიარა — ბიუჯეტი DWG-ს ემთხვევა")
    else:
        if failures:
            print("  ᲨᲔᲓᲔᲒᲘ: ✗ პრობლემები:")
            for f in failures:
                print(f"    • {f}")
        if warnings:
            print("  გაფრთხილება:")
            for w in warnings:
                print(f"    ⚠ {w}")
    print("="*68 + "\n")

    return len(failures) == 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("გამოყენება: python3 compare.py <dwg_ან_dxf> <excel.xlsx>")
        sys.exit(1)
    ok = compare(sys.argv[1], sys.argv[2])
    sys.exit(0 if ok else 1)

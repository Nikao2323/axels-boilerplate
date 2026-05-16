"""
Excel Report Generator
ქმნის formatted შედარების ანგარიშს compare.py-ის შედეგებიდან.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Colours ──────────────────────────────────────────────────────────────────
C_HEADER   = "1F3864"   # dark navy
C_SECTION  = "2F5496"   # mid blue
C_OK       = "E2EFDA"   # light green
C_WARN     = "FFF2CC"   # light yellow
C_FAIL     = "FCE4D6"   # light red
C_LABEL    = "D9E1F2"   # light blue-grey
C_WHITE    = "FFFFFF"
C_BORDER   = "8EA9C1"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write(ws, row, col, value, fill=None, font=None, align=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill    = fill
    if font:   cell.font    = font
    if align:  cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    cell.border = BORDER
    return cell


def _merge_write(ws, r1, c1, r2, c2, value, fill=None, font=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    cell.border = BORDER
    return cell


# ── Main sheet builder ────────────────────────────────────────────────────────

def build_report(all_results: list, output_path: str):
    """
    all_results: list of dicts returned by run_all.py
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "შეჯამება"
    _build_summary(ws_sum, all_results)

    # ── One sheet per building ────────────────────────────────────────────────
    for res in all_results:
        ws = wb.create_sheet(title=res["name"][:31])
        _build_detail(ws, res)

    wb.save(output_path)
    print(f"  ✓ Report saved: {output_path}")


def _build_summary(ws, all_results):
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # Title
    _merge_write(ws, 1, 1, 1, 9,
                 "DWG vs Excel — შედარების ანგარიში",
                 fill=_fill(C_HEADER),
                 font=_font(bold=True, color="FFFFFF", size=14),
                 align=_align("center"))
    _merge_write(ws, 2, 1, 2, 9,
                 f"გენერირდა: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                 fill=_fill(C_LABEL),
                 font=_font(color="444444", size=9),
                 align=_align("center"))

    # Column headers
    headers = ["შენობა", "სულ ფართი\nExcel m²", "სულ ფართი\nDWG m²",
               "ფართი\nსხვაობა", "კარკასი m²",
               "ბეტ. B-25 m³", "არმ. A500 t",
               "ქვაბ. ამოღ. m³", "სტატუსი"]
    row = 4
    for c, h in enumerate(headers, 1):
        _write(ws, row, c, h,
               fill=_fill(C_SECTION),
               font=_font(bold=True, color="FFFFFF", size=9),
               align=_align("center", wrap=True))
    ws.row_dimensions[row].height = 36

    # Data rows
    for i, res in enumerate(all_results):
        row = 5 + i
        af = res["area_floors"]
        br = res["budget_ratios"]

        total_xl  = af.get("სულ_excel", 0)
        total_dwg = af.get("სულ_dwg", 0)
        diff_pct  = af.get("სულ_diff_pct", 0)

        overall_ok = res["ok"]
        row_fill = _fill(C_OK) if overall_ok else _fill(C_FAIL)

        _write(ws, row, 1, res["name"],    fill=row_fill, font=_font(bold=True))
        _write(ws, row, 2, total_xl,       fill=row_fill, num_fmt='#,##0.0')
        _write(ws, row, 3, total_dwg,      fill=row_fill, num_fmt='#,##0.0')
        _write(ws, row, 4, diff_pct/100,   fill=_fill(C_OK) if abs(diff_pct) <= 10 else _fill(C_FAIL),
               num_fmt='0.0%', align=_align("center"))

        # ratio columns
        ratio_keys = ["karkasi", "betoni_b25", "armatura_a500", "qvabuli_amoreba"]
        for col_off, rk in enumerate(ratio_keys):
            info = br.get(rk, {})
            qty  = info.get("qty")
            ok   = info.get("ok")
            cf   = _fill(C_OK) if ok else (_fill(C_WARN) if info.get("warn_only") else _fill(C_FAIL))
            _write(ws, row, 5 + col_off,
                   qty if qty else "—",
                   fill=cf,
                   num_fmt='#,##0.00',
                   align=_align("right"))

        status_txt = "✓ OK" if overall_ok else "✗ შეამოწმე"
        _write(ws, row, 9, status_txt,
               fill=_fill(C_OK) if overall_ok else _fill(C_FAIL),
               font=_font(bold=True, color="276221" if overall_ok else "C00000"),
               align=_align("center"))

    # Totals row
    n = len(all_results)
    row = 5 + n
    ws.row_dimensions[row].height = 16
    _merge_write(ws, row, 1, row, 3, f"სულ: {n} შენობა",
                 fill=_fill(C_LABEL), font=_font(bold=True),
                 align=_align("center"))
    ok_count = sum(1 for r in all_results if r["ok"])
    _merge_write(ws, row, 4, row, 8,
                 f"გაიარა: {ok_count}/{n}",
                 fill=_fill(C_LABEL), font=_font(bold=True),
                 align=_align("center"))
    _write(ws, row, 9,
           "✓ ყველა" if ok_count == n else f"✗ {n-ok_count} პრობლემა",
           fill=_fill(C_OK) if ok_count == n else _fill(C_FAIL),
           font=_font(bold=True, color="276221" if ok_count == n else "C00000"),
           align=_align("center"))


def _build_detail(ws, res):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.row_dimensions[1].height = 28

    name = res["name"]
    _merge_write(ws, 1, 1, 1, 7, f"{name} — DWG vs Excel შედარება",
                 fill=_fill(C_HEADER),
                 font=_font(bold=True, color="FFFFFF", size=12),
                 align=_align("center"))

    row = 3

    # ── A: Floor areas ────────────────────────────────────────────────────
    _merge_write(ws, row, 1, row, 7, "A. სართულების ფართები",
                 fill=_fill(C_SECTION),
                 font=_font(bold=True, color="FFFFFF"),
                 align=_align("center"))
    row += 1

    hdrs = ["", "სართული", "Excel m²", "DWG m²", "სხვაობა %", "ლიმიტი", "სტატ."]
    for c, h in enumerate(hdrs, 1):
        _write(ws, row, c, h,
               fill=_fill(C_LABEL),
               font=_font(bold=True, size=9),
               align=_align("center"))
    row += 1

    for floor_name, fa in sorted(res["area_floors"].items(),
                                  key=lambda x: -x[1].get("excel", 0)
                                  if isinstance(x[1], dict) else 0):
        if not isinstance(fa, dict):
            continue
        excel_a = fa.get("excel", 0)
        dwg_a   = fa.get("dwg", 0)
        diff    = fa.get("diff_pct", 0)
        ok      = fa.get("ok", False)
        fill    = _fill(C_OK) if ok else _fill(C_FAIL)

        _write(ws, row, 1, "", fill=fill)
        _write(ws, row, 2, floor_name, fill=fill, font=_font(bold=(floor_name == "სულ")))
        _write(ws, row, 3, excel_a,    fill=fill, num_fmt='#,##0.0')
        _write(ws, row, 4, dwg_a if dwg_a else "—", fill=fill, num_fmt='#,##0.0')
        _write(ws, row, 5, diff/100 if diff else 0,  fill=fill, num_fmt='0.0%', align=_align("center"))
        _write(ws, row, 6, "±8%",      fill=fill, align=_align("center"))
        _write(ws, row, 7, "✓" if ok else "✗",
               fill=fill, font=_font(bold=True, color="276221" if ok else "C00000"),
               align=_align("center"))
        row += 1

    row += 1

    # ── B: Budget ratios ──────────────────────────────────────────────────
    _merge_write(ws, row, 1, row, 7, "B. ბიუჯეტის მოცულობები (raod. სვეტი)",
                 fill=_fill(C_SECTION),
                 font=_font(bold=True, color="FFFFFF"),
                 align=_align("center"))
    row += 1

    hdrs2 = ["", "პუნქტი", "ერთ.", "Excel raod.", "მოსალოდ. min", "მოსალოდ. max", "სტატ."]
    for c, h in enumerate(hdrs2, 1):
        _write(ws, row, c, h,
               fill=_fill(C_LABEL),
               font=_font(bold=True, size=9),
               align=_align("center"))
    row += 1

    for rk, info in res["budget_ratios"].items():
        ok       = info.get("ok", False)
        warn     = info.get("warn_only", False)
        if ok:
            fill = _fill(C_OK)
        elif warn:
            fill = _fill(C_WARN)
        else:
            fill = _fill(C_FAIL)

        _write(ws, row, 1, "", fill=fill)
        _write(ws, row, 2, info.get("label", rk), fill=fill, font=_font(bold=True))
        _write(ws, row, 3, info.get("unit", ""),  fill=fill, align=_align("center"))
        _write(ws, row, 4, info.get("qty") if info.get("qty") else "—",
               fill=fill, num_fmt='#,##0.00')
        _write(ws, row, 5, info.get("exp_min"), fill=fill, num_fmt='#,##0.0')
        _write(ws, row, 6, info.get("exp_max"), fill=fill, num_fmt='#,##0.0')
        tag = "✓" if ok else ("⚠" if warn else "✗")
        _write(ws, row, 7, tag,
               fill=fill,
               font=_font(bold=True,
                          color="276221" if ok else ("7F6000" if warn else "C00000")),
               align=_align("center"))
        row += 1

    row += 1

    # ── C: Full budget list ───────────────────────────────────────────────
    _merge_write(ws, row, 1, row, 7, "C. ბიუჯეტის სრული სია",
                 fill=_fill(C_SECTION),
                 font=_font(bold=True, color="FFFFFF"),
                 align=_align("center"))
    row += 1

    hdrs3 = ["", "სამუშაოს დასახელება", "", "", "ერთ.", "raod.", ""]
    for c, h in enumerate(hdrs3, 1):
        _write(ws, row, c, h,
               fill=_fill(C_LABEL),
               font=_font(bold=True, size=9),
               align=_align("center"))
    row += 1

    for item in res.get("full_budget", []):
        alt = _fill("F2F2F2") if row % 2 == 0 else _fill(C_WHITE)
        _write(ws, row, 1, "", fill=alt)
        _merge_write(ws, row, 2, row, 4, item["desc"], fill=alt, align=_align("left", wrap=True))
        _write(ws, row, 5, item["unit"],  fill=alt, align=_align("center"))
        _write(ws, row, 6, item["qty"],   fill=alt, num_fmt='#,##0.00')
        _write(ws, row, 7, "",            fill=alt)
        row += 1
